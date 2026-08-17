from __future__ import annotations

import argparse
import ast
import csv
import hashlib
import io
import json
import posixpath
import re
import subprocess
import sys
import zipfile
from datetime import date
from pathlib import Path, PurePosixPath
from typing import Any
from xml.etree import ElementTree

from project_utils import (
    build_directory,
    control_path,
    delivery_directory,
    ensure_safe_file,
    generated_directory,
    load_profile,
    numeric_limit,
    question_numbers,
)
from evidence_freeze import evidence_differences


INPUT_SUFFIXES = {".pdf", ".doc", ".docx", ".csv", ".tsv", ".xls", ".xlsx"}
SUPPORT_EXCLUDED_PARTS = {
    "__pycache__",
    ".pytest_cache",
    ".mypy_cache",
    ".ruff_cache",
    ".ipynb_checkpoints",
    "build",
}
NESTED_ARCHIVE_SUFFIXES = {".zip", ".rar", ".7z"}
TABULAR_RESULT_SUFFIXES = {".csv", ".tsv", ".xls", ".xlsx"}
FIGURE_SUFFIXES = {".pdf", ".png", ".jpg", ".jpeg", ".eps", ".svg"}
GENERIC_FIGURE_STEMS = {
    "chart",
    "figure",
    "final",
    "image",
    "output",
    "overview",
    "plot",
    "result",
    "summary",
}
RESULT_EXCEPTION_SCOPES = {"filename", "headers", "sheet_names", "encoding"}
HAN_RE = re.compile(r"[\u3400-\u4dbf\u4e00-\u9fff\uf900-\ufaff]")
SPREADSHEET_NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
OFFICE_REL_NS = (
    "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
)
PACKAGE_REL_NS = "{http://schemas.openxmlformats.org/package/2006/relationships}"
SCRATCH_DELIVERY_NAMES = {
    "audit_results.py",
    "build_support_zip.ps1",
    "cleanup_workspace.ps1",
    "make_pdf_contacts.py",
    "rebuild_support.ps1",
    "verify_results.py",
}


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def read_json(path: Path) -> dict[str, Any]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise ValueError("root must be an object")
    return payload


def centimeters(value: str) -> float | None:
    match = re.fullmatch(r"\s*([0-9]+(?:\.[0-9]+)?)\s*(cm|mm|in)\s*", value)
    if not match:
        return None
    number = float(match.group(1))
    return number if match.group(2) == "cm" else number / 10 if match.group(2) == "mm" else number * 2.54


def cumcm_margin_error(main_tex: str) -> str | None:
    match = re.search(r"\\usepackage\[([^]]+)]\{geometry}", main_tex)
    if not match:
        return "CUMCM paper must declare explicit geometry margins"
    options = {}
    for item in match.group(1).split(","):
        if "=" in item:
            key, value = item.split("=", 1)
            options[key.strip()] = value.strip()
    uniform = centimeters(options.get("margin", ""))
    if uniform is not None:
        return None if uniform >= 2.5 else f"CUMCM paper margin is {uniform:g} cm; minimum is 2.5 cm"
    values = [centimeters(options.get(key, "")) for key in ("top", "bottom", "left", "right")]
    if any(value is None for value in values):
        return "CUMCM paper must declare top, bottom, left, and right margins or one uniform margin"
    if min(value for value in values if value is not None) < 2.5:
        return "every CUMCM paper margin must be at least 2.5 cm"
    return None


def has_chinese(value: str) -> bool:
    return HAN_RE.search(value) is not None


def normalized_relative_path(value: str) -> PurePosixPath:
    normalized = PurePosixPath(value.replace("\\", "/"))
    if normalized.is_absolute() or not normalized.parts or ".." in normalized.parts:
        raise ValueError(f"path must be project-relative without traversal: {value}")
    return normalized


def first_symlink_component(root: Path, path: Path) -> Path | None:
    try:
        relative = path.relative_to(root)
    except ValueError:
        return path
    current = root
    for part in relative.parts:
        current /= part
        if current.is_symlink():
            return current
    return None


def fixed_schema_source_error(root: Path, value: str) -> str | None:
    try:
        relative = normalized_relative_path(value)
    except ValueError as error:
        return str(error)
    source = root.joinpath(*relative.parts)
    symlink = first_symlink_component(root, source)
    if symlink is not None:
        return f"fixed-schema source must not traverse a symbolic link: {value}"
    try:
        source.resolve().relative_to(root.resolve())
    except ValueError:
        return f"fixed-schema source escapes the project root: {value}"
    is_question_file = relative.parts[0] == "question"
    is_root_problem_file = (
        len(relative.parts) == 1 and source.suffix.lower() in INPUT_SUFFIXES
    )
    if not is_question_file and not is_root_problem_file:
        return (
            "fixed-schema source must be under question/ or be a supported "
            f"root-level problem input: {value}"
        )
    if not source.is_file():
        return f"fixed-schema source does not exist: {value}"
    return None


def delimited_header(
    path: Path, delimiter: str, allow_prescribed_encoding: bool = False
) -> tuple[bool, list[str]]:
    raw = path.read_bytes()
    has_utf8_bom = raw.startswith(b"\xef\xbb\xbf")
    encodings = (
        ("utf-8-sig", "gb18030", "utf-16")
        if allow_prescribed_encoding
        else ("utf-8-sig",)
    )
    decode_error: UnicodeDecodeError | None = None
    for encoding in encodings:
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError as error:
            decode_error = error
    else:
        assert decode_error is not None
        raise decode_error
    for row in csv.reader(io.StringIO(text), delimiter=delimiter):
        headers = [cell.strip() for cell in row]
        if any(headers):
            return has_utf8_bom, headers
    return has_utf8_bom, []


def xlsx_sheet_headers(path: Path) -> list[tuple[str, list[str]]]:
    """Read worksheet names and semantic header rows with an OOXML fallback."""
    try:
        import openpyxl
    except ImportError:
        pass
    else:
        try:
            workbook = openpyxl.load_workbook(
                path, read_only=False, data_only=False, keep_links=False
            )
        except (KeyError, OSError, ValueError, zipfile.BadZipFile):
            workbook = None
        if workbook is not None:
            sheets: list[tuple[str, list[str]]] = []
            try:
                for worksheet in workbook.worksheets:
                    candidate_rows: list[list[str]] = []
                    for row in worksheet.iter_rows(max_row=min(20, worksheet.max_row)):
                        values = [
                            "" if cell.value is None else str(cell.value).strip()
                            for cell in row
                        ]
                        if any(values):
                            candidate_rows.append(values)
                            if sum(bool(value) for value in values) >= 2:
                                break
                    headers = next(
                        (
                            values
                            for values in candidate_rows
                            if sum(bool(value) for value in values) >= 2
                        ),
                        candidate_rows[0] if candidate_rows else [],
                    )
                    sheets.append((worksheet.title.strip(), headers))
            finally:
                workbook.close()
            return sheets

    with zipfile.ZipFile(path) as workbook:
        shared_strings: list[str] = []
        if "xl/sharedStrings.xml" in workbook.namelist():
            shared_root = ElementTree.fromstring(workbook.read("xl/sharedStrings.xml"))
            for item in shared_root.iter(f"{SPREADSHEET_NS}si"):
                shared_strings.append(
                    "".join(node.text or "" for node in item.iter(f"{SPREADSHEET_NS}t"))
                )

        workbook_root = ElementTree.fromstring(workbook.read("xl/workbook.xml"))
        relationships_root = ElementTree.fromstring(
            workbook.read("xl/_rels/workbook.xml.rels")
        )
        targets = {
            relationship.attrib["Id"]: relationship.attrib["Target"]
            for relationship in relationships_root.iter(f"{PACKAGE_REL_NS}Relationship")
            if "Id" in relationship.attrib and "Target" in relationship.attrib
        }

        sheets: list[tuple[str, list[str]]] = []
        for sheet in workbook_root.iter(f"{SPREADSHEET_NS}sheet"):
            name = sheet.attrib.get("name", "").strip()
            relationship_id = sheet.attrib.get(f"{OFFICE_REL_NS}id")
            if not relationship_id or relationship_id not in targets:
                raise ValueError(f"worksheet relationship is missing for {name or 'unnamed sheet'}")
            target = targets[relationship_id].replace("\\", "/")
            member = target.lstrip("/")
            if not member.startswith("xl/"):
                member = posixpath.normpath(f"xl/{member}")
            if member.startswith("../") or not member.startswith("xl/"):
                raise ValueError(f"unsafe worksheet path in workbook: {target}")
            sheet_root = ElementTree.fromstring(workbook.read(member))
            candidate_rows: list[list[str]] = []
            for row in sheet_root.iter(f"{SPREADSHEET_NS}row"):
                values: list[str] = []
                for cell in row.findall(f"{SPREADSHEET_NS}c"):
                    cell_type = cell.attrib.get("t")
                    if cell_type == "inlineStr":
                        value = "".join(
                            node.text or ""
                            for node in cell.iter(f"{SPREADSHEET_NS}t")
                        )
                    else:
                        raw_value = cell.findtext(f"{SPREADSHEET_NS}v") or ""
                        if cell_type == "s" and raw_value:
                            try:
                                value = shared_strings[int(raw_value)]
                            except (IndexError, ValueError) as error:
                                raise ValueError(
                                    f"invalid shared-string index in worksheet {name}"
                                ) from error
                        else:
                            value = raw_value
                    values.append(value.strip())
                if any(values):
                    candidate_rows.append(values)
                    if sum(bool(value) for value in values) >= 2:
                        break
                if len(candidate_rows) >= 20:
                    break
            headers = next(
                (
                    values
                    for values in candidate_rows
                    if sum(bool(value) for value in values) >= 2
                ),
                candidate_rows[0] if candidate_rows else [],
            )
            sheets.append((name, headers))
        return sheets


def result_artifact_contract_errors(
    root: Path, profile: dict[str, object]
) -> list[str]:
    """Validate Chinese-facing result tables and narrow fixed-schema exceptions."""
    errors: list[str] = []
    flag_names = (
        "require_chinese_filenames",
        "require_chinese_headers",
        "require_chinese_sheet_names",
        "require_utf8_sig_csv",
    )
    flags: dict[str, bool] = {}
    for name in flag_names:
        value = profile.get(name, True)
        if value is not True:
            errors.append(
                f"result_artifacts.{name} must be true; use a provenance-backed "
                "fixed_schema_exceptions entry for a problem-prescribed schema"
            )
            value = True
        flags[name] = value

    raw_exceptions = profile.get("fixed_schema_exceptions", [])
    if not isinstance(raw_exceptions, list):
        errors.append("result_artifacts.fixed_schema_exceptions must be a list")
        raw_exceptions = []
    exceptions: dict[str, tuple[set[str], Path]] = {}
    for index, item in enumerate(raw_exceptions, start=1):
        location = f"result_artifacts.fixed_schema_exceptions[{index}]"
        if not isinstance(item, dict):
            errors.append(f"{location} must be an object")
            continue
        valid = True
        unknown = sorted(set(item) - {"path", "source", "reason", "allow"})
        if unknown:
            errors.append(f"{location} has unsupported fields: {', '.join(unknown)}")
            valid = False
        path_value = item.get("path")
        source_value = item.get("source")
        reason = item.get("reason")
        allow = item.get("allow")
        if not isinstance(path_value, str) or not path_value.strip():
            errors.append(f"{location}.path is required")
            valid = False
        else:
            try:
                relative = normalized_relative_path(path_value)
            except ValueError as error:
                errors.append(f"invalid {location}.path: {error}")
                valid = False
            else:
                in_results = relative.parts[0] == "results"
                in_compact_results = (
                    len(relative.parts) >= 2
                    and relative.parts[0] == "produce"
                    and relative.parts[1] == "results"
                )
                if not in_results and not in_compact_results:
                    errors.append(f"{location}.path must be under results/ or produce/results/")
                    valid = False
                if relative.suffix.lower() not in TABULAR_RESULT_SUFFIXES:
                    errors.append(f"{location}.path must identify a CSV/TSV/XLS/XLSX file")
                    valid = False
                target = root.joinpath(*relative.parts)
                if first_symlink_component(root, target) is not None or not target.is_file():
                    errors.append(f"{location}.path does not identify a regular file: {path_value}")
                    valid = False
        if not isinstance(source_value, str) or not source_value.strip():
            errors.append(f"{location}.source is required")
            valid = False
        else:
            source_error = fixed_schema_source_error(root, source_value)
            if source_error:
                errors.append(f"invalid {location}.source: {source_error}")
                valid = False
        if not isinstance(reason, str) or not reason.strip():
            errors.append(f"{location}.reason is required")
            valid = False
        if (
            not isinstance(allow, list)
            or not allow
            or any(not isinstance(scope, str) for scope in allow)
        ):
            errors.append(f"{location}.allow must be a non-empty string list")
            valid = False
            allowed_scopes: set[str] = set()
        else:
            allowed_scopes = set(allow)
            if len(allowed_scopes) != len(allow):
                errors.append(f"{location}.allow must not contain duplicates")
                valid = False
            invalid_scopes = sorted(allowed_scopes - RESULT_EXCEPTION_SCOPES)
            if invalid_scopes:
                errors.append(
                    f"{location}.allow has unsupported scopes: {', '.join(invalid_scopes)}"
                )
                valid = False
        if valid and isinstance(path_value, str):
            key = normalized_relative_path(path_value).as_posix().casefold()
            if key in exceptions:
                errors.append(f"duplicate fixed-schema exception path: {path_value}")
            else:
                assert isinstance(source_value, str)
                source_relative = normalized_relative_path(source_value)
                exceptions[key] = (
                    allowed_scopes,
                    root.joinpath(*source_relative.parts),
                )

    roots = (root / "results", root / "produce" / "results")
    for base in roots:
        if not base.exists():
            continue
        if first_symlink_component(root, base) is not None:
            errors.append(f"symbolic result directory is unsafe: {base.relative_to(root)}")
            continue
        for path in sorted(base.rglob("*")):
            relative = path.relative_to(root).as_posix()
            if path.is_symlink():
                errors.append(f"symbolic path under result artifacts is unsafe: {relative}")
                continue
            if path.suffix.lower() not in TABULAR_RESULT_SUFFIXES:
                continue
            if not path.is_file():
                errors.append(f"tabular result must be a regular file: {relative}")
                continue
            exception = exceptions.get(relative.casefold())
            allowances = exception[0] if exception else set()
            prescribed_source = exception[1] if exception else None
            suffix = path.suffix.lower()
            if flags["require_chinese_filenames"] and "filename" not in allowances:
                if not has_chinese(path.stem):
                    errors.append(
                        f"tabular result filename must contain a descriptive Chinese name: {relative}"
                    )

            if suffix in {".csv", ".tsv"}:
                irrelevant = allowances & {"sheet_names"}
                if irrelevant:
                    errors.append(
                        f"fixed-schema exception uses Excel-only scope for {relative}: sheet_names"
                    )
                try:
                    has_bom, headers = delimited_header(
                        path,
                        "\t" if suffix == ".tsv" else ",",
                        allow_prescribed_encoding="encoding" in allowances,
                    )
                except (OSError, UnicodeDecodeError, csv.Error) as error:
                    errors.append(f"could not read UTF-8 tabular result {relative}: {error}")
                    continue
                if (
                    flags["require_utf8_sig_csv"]
                    and "encoding" not in allowances
                    and not has_bom
                ):
                    errors.append(f"CSV/TSV result must use UTF-8-SIG: {relative}")
                if not headers:
                    errors.append(f"tabular result is missing a header row: {relative}")
                elif "headers" not in allowances:
                    if any(not header for header in headers):
                        errors.append(f"tabular result contains an empty header: {relative}")
                    duplicate_headers = sorted(
                        {header for header in headers if header and headers.count(header) > 1}
                    )
                    if duplicate_headers:
                        errors.append(
                            f"tabular result contains duplicate headers in {relative}: "
                            + ", ".join(duplicate_headers)
                        )
                    if flags["require_chinese_headers"]:
                        non_chinese = [header for header in headers if not has_chinese(header)]
                        if non_chinese:
                            errors.append(
                                f"tabular result headers must state their Chinese meaning in {relative}: "
                                + ", ".join(non_chinese)
                            )
            elif suffix == ".xlsx":
                if "encoding" in allowances:
                    errors.append(
                        f"fixed-schema exception uses CSV-only scope for {relative}: encoding"
                    )
                try:
                    sheets = xlsx_sheet_headers(path)
                except (KeyError, OSError, ValueError, zipfile.BadZipFile, ElementTree.ParseError) as error:
                    errors.append(f"could not inspect XLSX result {relative}: {error}")
                    continue
                if not sheets:
                    errors.append(f"XLSX result contains no worksheets: {relative}")
                for sheet_name, headers in sheets:
                    if (
                        flags["require_chinese_sheet_names"]
                        and "sheet_names" not in allowances
                        and not has_chinese(sheet_name)
                    ):
                        errors.append(
                            f"XLSX worksheet name must state its Chinese meaning in {relative}: {sheet_name or '<empty>'}"
                        )
                    if not headers:
                        errors.append(
                            f"XLSX worksheet is missing a header row in {relative}: {sheet_name or '<empty>'}"
                        )
                    elif flags["require_chinese_headers"] and "headers" not in allowances:
                        non_chinese = [header for header in headers if not has_chinese(header)]
                        if non_chinese:
                            errors.append(
                                f"XLSX headers must state their Chinese meaning in {relative}/{sheet_name}: "
                                + ", ".join(non_chinese)
                            )
            else:
                if "encoding" in allowances:
                    errors.append(
                        f"fixed-schema exception uses CSV-only scope for {relative}: encoding"
                    )
                required_scopes = set()
                if flags["require_chinese_headers"]:
                    required_scopes.add("headers")
                if flags["require_chinese_sheet_names"]:
                    required_scopes.add("sheet_names")
                missing_scopes = sorted(required_scopes - allowances)
                if missing_scopes:
                    errors.append(
                        f"legacy XLS result cannot be inspected safely; generate XLSX or register the prescribed schema scopes for {relative}: "
                        + ", ".join(missing_scopes)
                    )
            if prescribed_source is not None:
                errors.extend(
                    fixed_schema_congruence_errors(
                        path, prescribed_source, allowances, relative
                    )
                )
    return errors


def figure_artifact_contract_errors(
    root: Path, profile: dict[str, object]
) -> list[str]:
    """Require final figures to explain themselves in Chinese filenames."""
    errors: list[str] = []
    if profile.get("require_chinese_figure_filenames", True) is not True:
        errors.append(
            "result_artifacts.require_chinese_figure_filenames must be true"
        )
    figures = root / "figures"
    if not figures.is_dir() or figures.is_symlink():
        return errors
    formats_by_stem: dict[str, set[str]] = {}
    for path in figures.rglob("*"):
        if not path.is_file() or path.is_symlink():
            continue
        relative = path.relative_to(root).as_posix()
        if path.suffix.lower() not in FIGURE_SUFFIXES:
            errors.append(
                f"figures/ contains a non-figure artifact; move it to results/: {relative}"
            )
            continue
        stem = path.stem.strip()
        key = path.relative_to(figures).with_suffix("").as_posix().casefold()
        formats_by_stem.setdefault(key, set()).add(path.suffix.lower())
        if stem.lower() in GENERIC_FIGURE_STEMS or not has_chinese(stem):
            errors.append(
                f"figure filename must be descriptive Chinese, not a generic name: {relative}"
            )
    if profile.get("require_pdf_svg_figure_pair", True) is not True:
        errors.append("result_artifacts.require_pdf_svg_figure_pair must be true")
    else:
        for stem, suffixes in sorted(formats_by_stem.items()):
            missing = {".pdf", ".svg"} - suffixes
            if missing:
                errors.append(
                    f"final figure requires PDF and SVG vector pair: figures/{stem} "
                    f"(missing {', '.join(sorted(missing))})"
                )
    return errors


def fixed_schema_congruence_errors(
    target: Path, source: Path, allowances: set[str], relative: str
) -> list[str]:
    """Prove that every waived presentation field matches the cited source."""
    errors: list[str] = []
    target_suffix = target.suffix.lower()
    source_suffix = source.suffix.lower()
    if target_suffix != source_suffix:
        return [
            f"fixed-schema exception format differs from its source for {relative}: "
            f"{target_suffix} != {source_suffix}"
        ]
    # A problem attachment often prescribes an upload filename in its prose
    # rather than by the attachment's own basename. Provenance therefore
    # proves a filename waiver; structural fields below can be compared
    # byte-for-byte against the cited template.
    if target_suffix in {".csv", ".tsv"}:
        delimiter = "\t" if target_suffix == ".tsv" else ","
        try:
            target_bom, target_headers = delimited_header(
                target, delimiter, allow_prescribed_encoding=True
            )
            source_bom, source_headers = delimited_header(
                source, delimiter, allow_prescribed_encoding=True
            )
        except (OSError, UnicodeDecodeError, csv.Error) as error:
            return [f"could not compare fixed schema for {relative}: {error}"]
        if "headers" in allowances and target_headers != source_headers:
            errors.append(
                f"fixed-schema headers differ from the cited source for {relative}"
            )
        if "encoding" in allowances and target_bom != source_bom:
            errors.append(
                f"fixed-schema BOM/encoding marker differs from the cited source for {relative}"
            )
    elif target_suffix == ".xlsx":
        try:
            target_sheets = xlsx_sheet_headers(target)
            source_sheets = xlsx_sheet_headers(source)
        except (KeyError, OSError, ValueError, zipfile.BadZipFile, ElementTree.ParseError) as error:
            return [f"could not compare fixed XLSX schema for {relative}: {error}"]
        if "sheet_names" in allowances and [name for name, _ in target_sheets] != [
            name for name, _ in source_sheets
        ]:
            errors.append(
                f"fixed-schema worksheet names differ from the cited source for {relative}"
            )
        if "headers" in allowances and [headers for _, headers in target_sheets] != [
            headers for _, headers in source_sheets
        ]:
            errors.append(
                f"fixed-schema worksheet headers differ from the cited source for {relative}"
            )
    return errors


def rules_freshness_errors(profile: dict[str, object]) -> list[str]:
    try:
        checked_date = date.fromisoformat(str(profile.get("rules_checked_at")))
        expires_date = date.fromisoformat(str(profile.get("rules_expires_at")))
    except ValueError:
        return ["contest profile must contain ISO rules_checked_at and rules_expires_at dates"]
    errors: list[str] = []
    today = date.today()
    if checked_date > today:
        errors.append("contest rules_checked_at must not be in the future")
    if expires_date < today:
        errors.append(
            f"contest rule snapshot expired on {expires_date.isoformat()}; recheck official sources before finalization"
        )
    if expires_date < checked_date:
        errors.append("contest rules_expires_at must not precede rules_checked_at")
    if profile.get("profile") == "cumcm-2026":
        official_ai_source = (
            "https://www.mcm.edu.cn/html_cn/node/"
            "fef94648f2836ab6cc81586f4c38512b.html"
        )
        sources = profile.get("sources")
        if not isinstance(sources, list) or official_ai_source not in sources:
            errors.append(
                "cumcm-2026 profile must cite the official 2026 AI-use rule"
            )
        ai_profile = profile.get("ai")
        if not isinstance(ai_profile, dict):
            errors.append("cumcm-2026 profile requires an ai policy object")
        else:
            if ai_profile.get("policy") != "cumcm-2026-trial":
                errors.append(
                    "cumcm-2026 profile must use the official cumcm-2026-trial AI policy"
                )
            submission_intent = ai_profile.get("submission_intent")
            if submission_intent not in {"training", "formal"}:
                errors.append(
                    "cumcm-2026 ai.submission_intent must be training or formal"
                )
            if submission_intent == "formal":
                if ai_profile.get("core_modeling_led_by_team") is not True:
                    errors.append(
                        "formal CUMCM submission requires team-led core modeling"
                    )
                if ai_profile.get("manual_review_completed") is not True:
                    errors.append(
                        "formal CUMCM submission requires completed manual AI review"
                    )
            for field in (
                "declaration_required",
                "declaration_before_references",
                "details_pdf_required",
            ):
                if ai_profile.get(field) is not True:
                    errors.append(f"cumcm-2026 ai.{field} must be true")
    return errors


def ai_declaration_errors(
    root: Path,
    main_tex: Path,
    ai_profile: dict[str, object],
) -> list[str]:
    """Validate the official 2026 declaration text and placement."""
    errors: list[str] = []
    if not bool(ai_profile.get("declaration_required")):
        return errors
    declaration_source = delivery_directory(root) / "sections" / "ai-disclosure.tex"
    declaration_text = (
        declaration_source.read_text(encoding="utf-8", errors="replace")
        if declaration_source.is_file()
        else main_tex.read_text(encoding="utf-8", errors="replace")
        if main_tex.is_file()
        else ""
    )
    compact = re.sub(r"\s+", "", declaration_text)
    if bool(ai_profile.get("used")):
        required_prefix = "本参赛队在竞赛过程中使用了AI工具，主要用于"
        required_suffix = "，详细使用情况见支撑材料。"
        if required_prefix not in compact or required_suffix not in compact:
            errors.append(
                "used-AI paper must include the official 2026 AI declaration text"
            )
    elif "本参赛队在竞赛过程中未使用任何AI工具。" not in compact:
        errors.append(
            "unused-AI paper must include the exact official 2026 AI declaration"
        )

    if bool(ai_profile.get("declaration_before_references")):
        if not main_tex.is_file():
            errors.append("cannot verify AI declaration placement without happy/main.tex")
            return errors
        main = re.sub(
            r"\s+",
            "",
            main_tex.read_text(encoding="utf-8", errors="replace"),
        )
        declaration_positions = [
            position
            for marker in (
                r"\input{sections/ai-disclosure}",
                "AI工具使用声明",
            )
            if (position := main.find(marker)) >= 0
        ]
        reference_positions = [
            position
            for marker in (
                r"\input{sections/references}",
                r"\begin{thebibliography}",
                "参考文献",
            )
            if (position := main.find(marker)) >= 0
        ]
        if (
            not declaration_positions
            or not reference_positions
            or min(declaration_positions) >= min(reference_positions)
        ):
            errors.append("AI工具使用声明 must appear immediately before references")
    return errors


def expected_support_names(root: Path, ai_profile: dict[str, object], support_profile: dict[str, object]) -> set[str]:
    names: set[str] = set()
    for directory in ("code", "results", "figures"):
        base = root / directory
        if not base.is_dir() or base.is_symlink():
            continue
        for path in base.rglob("*"):
            if path.is_file() and not path.is_symlink() and not any(part in SUPPORT_EXCLUDED_PARTS for part in path.relative_to(root).parts):
                names.add(path.relative_to(root).as_posix())
    if bool(ai_profile.get("used")) and bool(ai_profile.get("details_pdf_required")):
        names.add("AI工具使用详情.pdf")
    ai_log = control_path(
        root,
        "ai-use-log.md",
        delivery_directory(root) / "ai-use-log.md",
    )
    if bool(support_profile.get("include_ai_log")) and ai_log.is_file():
        names.add("AI工具使用记录.md")
    return names


def load_evidence_yaml(path: Path) -> dict[str, Any]:
    try:
        import yaml
    except ImportError:
        return load_simple_evidence_yaml(path)
    try:
        payload = yaml.safe_load(path.read_text(encoding="utf-8"))
    except Exception as error:  # PyYAML exposes parser-specific subclasses.
        raise ValueError(str(error)) from error
    if not isinstance(payload, dict):
        raise ValueError("root must be a mapping")
    return payload


def load_simple_evidence_yaml(path: Path) -> dict[str, Any]:
    """Parse the bundled flat qN/claims schema when PyYAML is unavailable."""
    payload: dict[str, Any] = {}
    current_question: dict[str, Any] | None = None
    current_claim: dict[str, str] | None = None
    for number, raw in enumerate(path.read_text(encoding="utf-8").splitlines(), start=1):
        line = raw.split("#", 1)[0].rstrip()
        if not line.strip():
            continue
        question_match = re.fullmatch(r"(q[1-9]\d*):", line)
        if question_match:
            current_question = {"claims": []}
            payload[question_match.group(1)] = current_question
            current_claim = None
            continue
        stripped = line.strip()
        if stripped == "claims:" and current_question is not None:
            continue
        field_match = re.fullmatch(
            r"(?:-\s*)?([A-Za-z_][A-Za-z0-9_]*):\s*(.*?)\s*",
            stripped,
        )
        if not field_match or current_question is None:
            raise ValueError(
                f"unsupported evidence YAML syntax at line {number}: {raw}"
            )
        field, value = field_match.groups()
        if stripped.startswith("-"):
            current_claim = {}
            current_question["claims"].append(current_claim)
        if current_claim is None:
            raise ValueError(f"claim field outside a list item at line {number}")
        current_claim[field] = value.strip("\"'")
    return payload


def validate_evidence(root: Path, count: int, errors: list[str], warnings: list[str]) -> None:
    path = control_path(
        root,
        "evidence-map.yaml",
        delivery_directory(root) / "evidence-map.yaml",
    )
    if not path.is_file():
        errors.append("missing .cumcm/evidence-map.yaml")
        return
    try:
        payload = load_evidence_yaml(path)
    except (OSError, ValueError) as error:
        errors.append(f"invalid evidence-map.yaml: {error}")
        return
    valid_status = {"TODO", "verified"}
    seen_ids: set[str] = set()
    for number in range(1, count + 1):
        key = f"q{number}"
        block = payload.get(key)
        if not isinstance(block, dict) or not isinstance(block.get("claims"), list):
            errors.append(f"evidence map requires {key}.claims list")
            continue
        for index, claim in enumerate(block["claims"], start=1):
            location = f"{key}.claims[{index}]"
            if not isinstance(claim, dict):
                errors.append(f"{location} must be a mapping")
                continue
            claim_id = claim.get("id")
            if not isinstance(claim_id, str) or not claim_id.strip():
                errors.append(f"{location}.id is required")
            elif claim_id in seen_ids:
                errors.append(f"duplicate evidence id: {claim_id}")
            else:
                seen_ids.add(claim_id)
            for field, prefix in (("source", "results"), ("paper_section", "happy")):
                value = claim.get(field)
                if not isinstance(value, str) or not value:
                    errors.append(f"{location}.{field} is required")
                    continue
                candidate = root / value
                try:
                    resolved = ensure_safe_file(root, candidate, {prefix})
                except ValueError as error:
                    errors.append(f"unsafe {location}.{field}: {error}")
                    continue
                relative = resolved.relative_to(root).as_posix()
                if field == "paper_section" and not (
                    relative == "happy/main.tex"
                    or relative.startswith("happy/sections/")
                ):
                    errors.append(
                        f"{location}.paper_section must be happy/main.tex "
                        "or a legacy happy/sections file"
                    )
                if not resolved.is_file():
                    errors.append(f"evidence path does not exist: {value}")
            status = claim.get("status")
            if status not in valid_status:
                errors.append(f"{location}.status must be TODO or verified")
            elif status == "TODO":
                warnings.append(f"unresolved evidence claim: {claim_id or location}")


def code_architecture_errors(root: Path, count: int) -> list[str]:
    """Reject incomplete qN implementations that cannot support a final paper."""
    errors: list[str] = []
    for number in range(1, count + 1):
        question_dir = root / "code" / f"q{number}"
        if not question_dir.is_dir():
            continue
        modules = sorted(
            path
            for path in question_dir.glob("*.py")
            if path.name != "__init__.py" and path.is_file()
        )
        names = {path.name for path in modules}
        if names == {"main.py"}:
            errors.append(
                f"q{number} contains only main.py; keep it orchestral and add "
                "responsibility-named modules for the actual model and validation"
            )
        for module_path in modules:
            module_text = module_path.read_text(encoding="utf-8", errors="replace")
            if "NotImplementedError" in module_text or "TODO[" in module_text:
                errors.append(
                    f"{module_path.relative_to(root).as_posix()} is still an unimplemented scaffold"
                )
    return errors


def code_architecture_warnings(root: Path, count: int) -> list[str]:
    """Find review-hostile Python layouts without prescribing one fixed architecture."""
    warnings: list[str] = []
    for number in range(1, count + 1):
        question_dir = root / "code" / f"q{number}"
        if not question_dir.is_dir():
            continue
        modules = sorted(
            path
            for path in question_dir.glob("*.py")
            if path.name != "__init__.py" and path.is_file()
        )
        names = {path.name for path in modules}
        if names == {"main.py", "model.py", "output.py"}:
            warnings.append(
                f"q{number} uses the generic main.py + model.py + output.py trio; "
                "split substantial logic by actual responsibility"
            )
        for path in modules:
            lines = path.read_text(encoding="utf-8", errors="replace").splitlines()
            code_lines = [line for line in lines if line.strip()]
            comments = [line for line in lines if line.lstrip().startswith("#")]
            if len(code_lines) >= 80 and not comments:
                warnings.append(
                    f"{path.relative_to(root).as_posix()} has {len(code_lines)} nonblank "
                    "lines but no review-oriented # comments"
                )
        main_path = question_dir / "main.py"
        if main_path.is_file():
            main_text = main_path.read_text(encoding="utf-8", errors="replace")
            main_lines = [
                line
                for line in main_text.splitlines()
                if line.strip()
            ]
            if len(main_lines) >= 150:
                warnings.append(
                    f"code/q{number}/main.py has {len(main_lines)} nonblank lines; "
                    "keep the entry point orchestral"
                )
    return warnings


def delivery_hygiene_warnings(root: Path) -> list[str]:
    """Flag authoring leftovers that should not enter a compact submission."""
    warnings: list[str] = []
    bases = [root / directory for directory in ("code", "results", "figures")]
    bases.extend(root / "produce" / directory for directory in ("code", "results", "figures"))
    for base in bases:
        if not base.is_dir() or base.is_symlink():
            continue
        for path in base.rglob("*"):
            if not path.is_file() or path.is_symlink():
                continue
            relative = path.relative_to(root).as_posix()
            if path.suffix.lower() in NESTED_ARCHIVE_SUFFIXES:
                warnings.append(f"nested project archive should be removed before handoff: {relative}")
            if path.name.lower() in SCRATCH_DELIVERY_NAMES:
                warnings.append(f"authoring-only helper should not be delivered as model code: {relative}")
    for directory in ("produce", "code", "results", "figures"):
        if not (root / directory).is_dir():
            continue
        for suffix in NESTED_ARCHIVE_SUFFIXES:
            duplicate = root / f"{directory}{suffix}"
            if duplicate.is_file() and not duplicate.is_symlink():
                warnings.append(
                    f"duplicate directory archive should be removed before handoff: {duplicate.name}"
                )
    temporary = root / "tmp"
    if temporary.is_dir() and any(temporary.iterdir()):
        warnings.append("non-empty tmp/ directory should be cleaned before handoff")
    manifest = control_path(
        root,
        "project.json",
        delivery_directory(root) / "question-count.json",
    )
    layout = ""
    if manifest.is_file():
        try:
            layout = str(read_json(manifest).get("layout", ""))
        except (OSError, ValueError, json.JSONDecodeError):
            pass
    if layout == "compact-v2":
        for directory in (
            delivery_directory(root) / "sections",
            delivery_directory(root) / "build",
        ):
            if directory.is_dir():
                warnings.append(
                    f"compact layout must not expose authoring directory: "
                    f"{directory.relative_to(root).as_posix()}/"
                )
        legacy_controls = (
            "contest-profile.json",
            "question-count.json",
            "evidence-map.yaml",
            "input-inventory.json",
            "question-fingerprint.json",
            "evidence-freeze.json",
            "ai-use-log.md",
        )
        for name in legacy_controls:
            if (delivery_directory(root) / name).exists():
                warnings.append(
                    f"move internal workflow state out of happy/: happy/{name}"
                )
        if (
            (root / "code" / "finalize.py").exists()
            or (control_path(root, "finalize.py")).exists()
        ):
            warnings.append(
                "project-local finalize.py is redundant; call the active "
                "cumcm-draft/scripts/finalize_project.py directly"
            )
        allowed_happy_files = {
            "AI工具使用详情.pdf",
            "AI工具使用详情.tex",
            "main.pdf",
            "main.tex",
            "支撑材料.zip",
        }
        compile_suffixes = {
            ".aux",
            ".bbl",
            ".bcf",
            ".blg",
            ".fdb_latexmk",
            ".fls",
            ".log",
            ".out",
            ".run.xml",
            ".synctex.gz",
            ".toc",
            ".xdv",
        }
        happy = delivery_directory(root)
        if happy.is_dir():
            for path in happy.iterdir():
                if not path.is_file() or path.name in allowed_happy_files:
                    continue
                if any(path.name.endswith(suffix) for suffix in compile_suffixes):
                    continue
                warnings.append(
                    f"happy/ contains a non-delivery internal artifact: "
                    f"{path.relative_to(root).as_posix()}"
                )
    for summary in root.glob("results/q*/summary.json"):
        if summary.is_file():
            warnings.append(
                f"generic control summary does not belong in final results: "
                f"{summary.relative_to(root).as_posix()}"
            )
    return warnings


def main() -> None:
    parser = argparse.ArgumentParser(description="Audit a modeling project and active contest profile.")
    parser.add_argument("project_root", type=Path)
    parser.add_argument("--strict", action="store_true")
    parser.add_argument("--run-code", action="store_true")
    args = parser.parse_args()
    root = args.project_root.resolve()
    errors: list[str] = []
    warnings: list[str] = []
    try:
        profile = load_profile(root)
        paper_profile = profile["paper"]
        support_profile = profile["support"]
        result_artifact_profile = profile["result_artifacts"]
        ai_profile = profile["ai"]
        assert (
            isinstance(paper_profile, dict)
            and isinstance(support_profile, dict)
            and isinstance(result_artifact_profile, dict)
            and isinstance(ai_profile, dict)
        )
        pdf_limit = numeric_limit(paper_profile, "max_pdf_mb")
        page_limit = numeric_limit(paper_profile, "max_body_pages")
        archive_limit = numeric_limit(support_profile, "max_archive_mb")
    except (ValueError, TypeError, AssertionError) as error:
        errors.append(str(error))
        profile = {}
        paper_profile = support_profile = ai_profile = {}
        result_artifact_profile = None
        pdf_limit = page_limit = archive_limit = None

    for name in ("question", "code", "results", "figures", "happy"):
        path = root / name
        if not path.is_dir():
            errors.append(f"missing top-level directory: {name}/")
        elif path.is_symlink():
            errors.append(f"symbolic top-level directory is unsafe: {name}/")

    count_path = control_path(
        root,
        "project.json",
        delivery_directory(root) / "question-count.json",
    )
    count = 0
    if not count_path.is_file():
        errors.append("missing .cumcm/project.json")
    else:
        try:
            count = int(read_json(count_path).get("questions", 0))
        except (OSError, ValueError, json.JSONDecodeError) as error:
            errors.append(f"invalid project manifest: {error}")
        if not 1 <= count <= 20:
            errors.append("project manifest must declare 1 to 20 questions")
    discovered = question_numbers(root)
    if discovered and max(discovered) > count:
        errors.append(f"project manifest omits discovered subproblem q{max(discovered)}")

    if args.run_code:
        entry = root / "code" / "run_all.py"
        if not entry.is_file():
            errors.append("missing code/run_all.py")
        else:
            completed = subprocess.run([sys.executable, str(entry)], cwd=root / "code", stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True, encoding="utf-8", errors="replace", check=False)
            if completed.returncode:
                errors.append("code/run_all.py failed:\n" + "\n".join(completed.stdout.splitlines()[-25:]))
            else:
                print("[OK] code/run_all.py completed")

    for number in range(1, count + 1):
        name = f"q{number}"
        required = (
            f"code/{name}",
            f"code/{name}/main.py",
            f"results/{name}",
            f"figures/{name}",
        )
        for relative in required:
            if not (root / relative).exists():
                errors.append(f"missing subproblem artifact: {relative}")
        main_source = delivery_directory(root) / "main.tex"
        legacy_section = delivery_directory(root) / "sections" / f"{name}.tex"
        if not legacy_section.is_file() and main_source.is_file():
            main_content = main_source.read_text(encoding="utf-8", errors="replace")
            if not re.search(rf"\\section\{{问题{number}(?:[：:]|\}})", main_content):
                errors.append(f"happy/main.tex is missing the problem {number} section")

    inventory_path = control_path(
        root,
        "input-inventory.json",
        delivery_directory(root) / "input-inventory.json",
    )
    if not inventory_path.is_file():
        warnings.append("input inventory is missing; run inspect_inputs.py")
    else:
        try:
            inventory = read_json(inventory_path)
            unresolved = [item.get("path", "unknown") for item in inventory.get("files", []) if isinstance(item, dict) and item.get("status") != "ok" and not item.get("details", {}).get("blank")]
            if unresolved:
                warnings.append("unresolved input inspection: " + ", ".join(map(str, unresolved)))
        except (OSError, ValueError, json.JSONDecodeError) as error:
            errors.append(f"invalid input inventory: {error}")

    baseline = control_path(
        root,
        "question-fingerprint.json",
        delivery_directory(root) / "question-fingerprint.json",
    )
    if not baseline.is_file():
        warnings.append("question fingerprint baseline is missing; run inspect_inputs.py")
    else:
        try:
            expected = {str(item["path"]): str(item["sha256"]) for item in read_json(baseline).get("files", [])}
            candidates = []
            question = root / "question"
            if question.is_dir():
                candidates.extend(path for path in question.rglob("*") if path.is_file())
            candidates.extend(path for path in root.iterdir() if path.is_file() and path.suffix.lower() in INPUT_SUFFIXES)
            if any(path.is_symlink() for path in candidates):
                raise ValueError("symbolic input paths are not allowed")
            actual = {path.relative_to(root).as_posix(): sha256(path) for path in sorted(set(candidates))}
            if actual != expected:
                errors.append("problem inputs changed after their fingerprint baseline was recorded")
        except (KeyError, OSError, ValueError, json.JSONDecodeError) as error:
            errors.append(f"invalid question fingerprint: {error}")

    for source in sorted((root / "code").rglob("*.py")):
        if not source.is_file() or source.is_symlink():
            continue
        try:
            ast.parse(
                source.read_text(encoding="utf-8"),
                filename=source.relative_to(root).as_posix(),
            )
        except (OSError, SyntaxError, UnicodeDecodeError) as error:
            errors.append(
                f"Python syntax failed for {source.relative_to(root).as_posix()}: {error}"
            )
    errors.extend(code_architecture_errors(root, count))
    warnings.extend(code_architecture_warnings(root, count))
    warnings.extend(delivery_hygiene_warnings(root))
    if result_artifact_profile is not None:
        errors.extend(result_artifact_contract_errors(root, result_artifact_profile))
        errors.extend(figure_artifact_contract_errors(root, result_artifact_profile))
    if profile:
        errors.extend(rules_freshness_errors(profile))
    validate_evidence(root, count, errors, warnings)
    try:
        frozen_differences = evidence_differences(root)
        if frozen_differences:
            errors.extend(frozen_differences)
    except (KeyError, OSError, TypeError, ValueError, json.JSONDecodeError) as error:
        errors.append(f"invalid evidence freeze: {error}")

    tex_files = [
        path
        for base in (delivery_directory(root), generated_directory(root))
        if base.is_dir()
        for path in base.rglob("*.tex")
        if path.is_file() and not path.is_symlink()
    ]
    tex = "\n".join(path.read_text(encoding="utf-8", errors="replace") for path in tex_files)
    todo_count = tex.count(r"\TODO{") + len(re.findall(r"TODO\[", tex))
    if todo_count:
        warnings.append(f"paper contains {todo_count} visible TODO marker(s)")
    if re.search(r"[A-Za-z]:[\\/]", tex):
        errors.append("paper contains an absolute Windows path")
    if re.search(r"(姓名|学号|学校|指导教师)\s*[:：]\s*\S+", tex):
        warnings.append("paper may contain identifying information")
    main_tex = delivery_directory(root) / "main.tex"
    if main_tex.is_file():
        content = main_tex.read_text(encoding="utf-8", errors="replace")
        if profile.get("profile") == "cumcm-2026":
            margin_error = cumcm_margin_error(content)
            if margin_error:
                errors.append(margin_error)
            if r"\tableofcontents" in content:
                errors.append("CUMCM paper must not contain a table of contents")
    if isinstance(ai_profile, dict):
        errors.extend(ai_declaration_errors(root, main_tex, ai_profile))

    if bool(paper_profile.get("include_support_file_list")):
        support_tex = generated_directory(root) / "support-files.tex"
        if not support_tex.is_file():
            support_tex = delivery_directory(root) / "sections" / "support-files.tex"
        if not support_tex.is_file() or "支撑材料文件列表" not in support_tex.read_text(encoding="utf-8", errors="replace"):
            errors.append("CUMCM appendix is missing the required support-material file list")
    if bool(paper_profile.get("include_source_appendix")):
        source_tex = generated_directory(root) / "source-code.tex"
        if not source_tex.is_file():
            source_tex = delivery_directory(root) / "sections" / "source-code.tex"
        if not source_tex.is_file() or r"\lstinputlisting" not in source_tex.read_text(encoding="utf-8", errors="replace"):
            errors.append("CUMCM appendix is missing complete runnable source listings")

    pdf = delivery_directory(root) / "main.pdf"
    if not pdf.is_file():
        errors.append("happy/main.pdf is missing; run build_paper.py")
    else:
        if pdf_limit is not None and pdf.stat().st_size > pdf_limit * 1024 * 1024:
            errors.append(f"happy/main.pdf exceeds active {pdf_limit:g} MB limit")
        if any(path.stat().st_mtime > pdf.stat().st_mtime for path in tex_files):
            warnings.append("happy/main.pdf is older than at least one TeX source")
        if profile.get("profile") == "cumcm-2026":
            try:
                from pypdf import PdfReader
                from pypdf.errors import PdfReadError
            except ImportError as error:
                if "摘要" not in tex or "关键词" not in tex:
                    errors.append(f"could not verify CUMCM first-page content without pypdf: {error}")
            else:
                try:
                    first_page = (PdfReader(pdf).pages[0].extract_text() or "").strip()
                    if "摘要" not in first_page or "关键词" not in first_page:
                        errors.append("CUMCM electronic paper first page must contain the abstract and keywords")
                    if "目录" in first_page:
                        errors.append("CUMCM electronic paper first page must not be a table of contents")
                except (IndexError, OSError, PdfReadError, ValueError) as error:
                    warnings.append(f"could not verify CUMCM first-page content: {error}")

    archive = delivery_directory(root) / "支撑材料.zip"
    if not archive.is_file():
        warnings.append("支撑材料.zip is missing; run package_support.py")
    else:
        if archive_limit is not None and archive.stat().st_size > archive_limit * 1024 * 1024:
            errors.append(
                f"{archive.name} exceeds active {archive_limit:g} MB limit"
            )
        try:
            with zipfile.ZipFile(archive) as bundle:
                names = [name for name in bundle.namelist() if not name.endswith("/")]
            unsafe = [name for name in names if PurePosixPath(name.replace("\\", "/")).is_absolute() or ".." in PurePosixPath(name.replace("\\", "/")).parts or re.match(r"^[A-Za-z]:", name)]
            if unsafe:
                errors.append("support archive contains unsafe absolute/traversal paths")
            if any(name.replace("\\", "/").lower().startswith("question/") for name in names):
                errors.append("support archive contains immutable question input files")
            if any(name.endswith("ai-use-log.md") or name == "AI工具使用记录.md" for name in names) and not bool(support_profile.get("include_ai_log")):
                errors.append("support archive leaks the internal AI use log contrary to profile")
            expected_ai_name = "AI工具使用详情.pdf"
            if bool(ai_profile.get("used")) and bool(ai_profile.get("details_pdf_required")) and expected_ai_name not in names:
                errors.append("support archive is missing required AI工具使用详情.pdf")
            expected_names = expected_support_names(root, ai_profile, support_profile)
            actual_names = set(names)
            missing_names = sorted(expected_names - actual_names)
            extra_names = sorted(actual_names - expected_names)
            if missing_names:
                errors.append("support archive is missing listed project files: " + ", ".join(missing_names[:12]))
            if extra_names:
                errors.append("support archive contains unlisted project files: " + ", ".join(extra_names[:12]))
        except zipfile.BadZipFile:
            errors.append(f"{archive.relative_to(root)} is not a valid ZIP archive")

    aux = build_directory(root) / "main.aux"
    if not aux.is_file():
        aux = delivery_directory(root) / "build" / "main.aux"
    if page_limit is not None and aux.is_file():
        match = re.search(r"\\newlabel\{body-end\}\{\{.*?\}\{(\d+)\}", aux.read_text(encoding="utf-8", errors="replace"))
        if match:
            body_pages = max(int(match.group(1)) - 1, 0)
            if body_pages > page_limit:
                errors.append(f"paper body has {body_pages} pages; active limit is {page_limit:g}")
            else:
                print(f"  Body page audit: {body_pages}/{page_limit:g}")
        else:
            warnings.append("could not resolve body-end page label")

    if bool(ai_profile.get("used")):
        disclosure = control_path(
            root,
            "ai-use-log.md",
            delivery_directory(root) / "ai-use-log.md",
        )
        if not disclosure.is_file():
            warnings.append("internal AI use log is missing")
        if bool(ai_profile.get("details_pdf_required")) and not (delivery_directory(root) / "AI工具使用详情.pdf").is_file():
            errors.append("AI tool disclosure PDF is required but missing")
        details_source = delivery_directory(root) / "AI工具使用详情.tex"
        if bool(ai_profile.get("details_pdf_required")):
            details_text = (
                details_source.read_text(encoding="utf-8", errors="replace")
                if details_source.is_file()
                else ""
            )
            required_sections = ("主要提示方式与过程", "采纳、人工修改与核验")
            if not details_source.is_file() or any(
                section not in details_text for section in required_sections
            ):
                errors.append(
                    "AI details source must cover prompting/usage process and adoption/manual verification"
                )

    print(f"[RUN] Validation summary: {len(errors)} error(s), {len(warnings)} warning(s)")
    for message in errors:
        print(f"[ERROR] {message}")
    for message in warnings:
        print(f"[WARN] {message}")
    if errors or (args.strict and warnings):
        raise SystemExit(1)
    print("[OK] Structure, provenance, evidence, syntax, profile, and delivery checks passed")


if __name__ == "__main__":
    main()
