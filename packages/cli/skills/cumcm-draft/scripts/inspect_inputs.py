from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import shutil
import subprocess
import zipfile
from pathlib import Path

from project_utils import control_directory
from script_runtime import configure_utf8_output
from typing import Any
from xml.etree import ElementTree


SUPPORTED = {".pdf", ".doc", ".docx", ".csv", ".tsv", ".xls", ".xlsx"}


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def result(status: str, **details: Any) -> dict[str, Any]:
    return {"status": status, **details}


def json_default(value: Any) -> str:
    """把 Excel 日期等预览值稳定转换为可审计文本。"""
    isoformat = getattr(value, "isoformat", None)
    if callable(isoformat):
        return str(isoformat())
    return str(value)


def inspect_delimited(path: Path) -> dict[str, Any]:
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    for encoding in ("utf-8-sig", "gb18030"):
        rows = columns = nonblank = 0
        preview: list[list[str]] = []
        try:
            with path.open("r", encoding=encoding, newline="") as stream:
                for row in csv.reader(stream, delimiter=delimiter):
                    rows += 1
                    columns = max(columns, len(row))
                    nonblank += sum(bool(cell.strip()) for cell in row)
                    if len(preview) < 20:
                        preview.append(row[:20])
            return result("ok", rows=rows, columns=columns, blank=nonblank == 0, encoding=encoding, preview=preview)
        except UnicodeDecodeError:
            continue
    return result("manual_review", reason="text encoding could not be resolved")


def inspect_xlsx(path: Path) -> dict[str, Any]:
    try:
        import openpyxl
    except ImportError:
        return inspect_xlsx_xml(path)
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=False)
    sheets: list[dict[str, Any]] = []
    try:
        for sheet in workbook.worksheets:
            nonblank = sum(cell.value not in (None, "") for row in sheet.iter_rows() for cell in row)
            preview = [[cell.value for cell in row[:20]] for row in list(sheet.iter_rows(max_row=20))[:20]]
            sheets.append({"name": sheet.title, "rows": sheet.max_row, "columns": sheet.max_column, "nonblank_cells": nonblank, "blank": nonblank == 0, "preview": preview})
    finally:
        workbook.close()
    return result("ok", sheets=sheets, blank=all(item["blank"] for item in sheets))


def inspect_xlsx_xml(path: Path) -> dict[str, Any]:
    try:
        with zipfile.ZipFile(path) as archive:
            workbook = ElementTree.fromstring(archive.read("xl/workbook.xml"))
            namespace = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
            names = [node.attrib.get("name", "unknown") for node in workbook.findall("x:sheets/x:sheet", namespace)]
            worksheets = sorted(name for name in archive.namelist() if name.startswith("xl/worksheets/sheet") and name.endswith(".xml"))
            sheets = []
            for index, member in enumerate(worksheets):
                sheet_root = ElementTree.fromstring(archive.read(member))
                cells = sheet_root.findall(".//x:c", namespace)
                nonblank = sum(cell.find("x:f", namespace) is not None or cell.find("x:v", namespace) is not None or bool(cell.findall(".//x:t", namespace)) for cell in cells)
                sheets.append({"name": names[index] if index < len(names) else member, "nonblank_cells": nonblank, "blank": nonblank == 0})
        return result("ok", sheets=sheets, blank=all(item["blank"] for item in sheets), inspection="OOXML fallback")
    except (KeyError, zipfile.BadZipFile, ElementTree.ParseError) as error:
        return result("manual_review", reason=f"invalid XLSX container: {error}")


def inspect_xls(path: Path) -> dict[str, Any]:
    try:
        import xlrd
    except ImportError:
        return result("dependency_missing", dependency="xlrd", reason="install code/requirements.txt")
    try:
        workbook = xlrd.open_workbook(path, on_demand=True)
        sheets = []
        for sheet in workbook.sheets():
            nonblank = sum(bool(str(sheet.cell_value(row, col)).strip()) for row in range(sheet.nrows) for col in range(sheet.ncols))
            sheets.append({"name": sheet.name, "rows": sheet.nrows, "columns": sheet.ncols, "nonblank_cells": nonblank, "blank": nonblank == 0})
        workbook.release_resources()
        return result("ok", sheets=sheets, blank=all(item["blank"] for item in sheets))
    except (OSError, xlrd.XLRDError) as error:
        return result("manual_review", reason=f"legacy workbook could not be read: {error}")


def inspect_docx(path: Path) -> dict[str, Any]:
    try:
        with zipfile.ZipFile(path) as archive:
            root = ElementTree.fromstring(archive.read("word/document.xml"))
    except (KeyError, zipfile.BadZipFile, ElementTree.ParseError) as error:
        return result("manual_review", reason=f"invalid DOCX container: {error}")
    text = "".join(node.text or "" for node in root.iter() if node.tag.endswith("}t")).strip()
    return result("ok", characters=len(text), blank=not text, text_excerpt=text[:50000])


def inspect_doc(path: Path) -> dict[str, Any]:
    tool = shutil.which("antiword")
    if not tool:
        return result("dependency_missing", dependency="antiword or LibreOffice", reason="legacy .doc requires conversion or extraction")
    completed = subprocess.run([tool, str(path)], stdout=subprocess.PIPE, stderr=subprocess.PIPE, check=False)
    text = completed.stdout.decode("utf-8", errors="replace").strip()
    return result("ok" if completed.returncode == 0 else "manual_review", characters=len(text), blank=not text, text_excerpt=text[:50000])


def inspect_pdf(path: Path) -> dict[str, Any]:
    try:
        from pypdf import PdfReader
    except ImportError:
        PdfReader = None
    if PdfReader is not None:
        try:
            reader = PdfReader(path)
            extracted = "".join((page.extract_text() or "") for page in reader.pages[: min(5, len(reader.pages))]).strip()
        except Exception as error:  # pypdf exposes several parser-specific exceptions.
            return result("manual_review", reason=f"PDF extraction failed: {error}")
        likely_scanned = len(extracted) < max(80, min(300, len(reader.pages) * 30))
        return result("requires_ocr" if likely_scanned else "ok", backend="pypdf", pages=len(reader.pages), sampled_characters=len(extracted), likely_scanned=likely_scanned, text_excerpt=extracted[:50000])

    pdf_module = None
    try:
        import pymupdf as pdf_module
    except ImportError:
        try:
            import fitz as pdf_module
        except ImportError:
            pass
    if pdf_module is not None:
        try:
            document = pdf_module.open(str(path))
            extracted = "".join((document[index].get_text() or "") for index in range(min(5, len(document)))).strip()
            likely_scanned = len(extracted) < max(80, min(300, len(document) * 30))
            return result("requires_ocr" if likely_scanned else "ok", backend="pymupdf", pages=len(document), sampled_characters=len(extracted), likely_scanned=likely_scanned, text_excerpt=extracted[:50000])
        except Exception as error:
            return result("manual_review", reason=f"PDF extraction failed: {error}")

    pdfinfo = shutil.which("pdfinfo.exe") or shutil.which("pdfinfo")
    pdftotext = shutil.which("pdftotext.exe") or shutil.which("pdftotext")
    if pdfinfo:
        completed = subprocess.run([pdfinfo, str(path)], text=True, encoding="utf-8", errors="replace", stdout=subprocess.PIPE, stderr=subprocess.DEVNULL, check=False)
        match = re.search(r"^Pages:\s+(\d+)", completed.stdout, re.MULTILINE)
        pages = int(match.group(1)) if match else None
        if completed.returncode == 0 and pages and pdftotext:
            text_process = subprocess.run([pdftotext, "-f", "1", "-l", str(min(5, pages)), str(path), "-"], text=True, encoding="utf-8", errors="replace", stdout=subprocess.PIPE, stderr=subprocess.DEVNULL, check=False)
            if text_process.returncode == 0:
                extracted = text_process.stdout.strip()
                likely_scanned = len(extracted) < max(80, min(300, pages * 30))
                return result("requires_ocr" if likely_scanned else "ok", backend="poppler", pages=pages, sampled_characters=len(extracted), likely_scanned=likely_scanned, text_excerpt=extracted[:50000])
        return result("manual_review", pages=pages, reason="text/scanned status not inspected")
    return result("dependency_missing", dependency="pypdf, pymupdf, or Poppler", reason="cannot inspect PDF text/scans")


def inspect_file(path: Path) -> dict[str, Any]:
    suffix = path.suffix.lower()
    if suffix in {".csv", ".tsv"}:
        return inspect_delimited(path)
    if suffix == ".xlsx":
        return inspect_xlsx(path)
    if suffix == ".xls":
        return inspect_xls(path)
    if suffix == ".docx":
        return inspect_docx(path)
    if suffix == ".doc":
        return inspect_doc(path)
    if suffix == ".pdf":
        return inspect_pdf(path)
    return result("unsupported", reason=f"unsupported extension {suffix or '[none]'}")


def discover_input_files(root: Path) -> list[Path]:
    question = root / "question"
    if question.exists() and (not question.is_dir() or question.is_symlink()):
        raise SystemExit(f"Unsafe question path: {question}")
    if question.is_dir():
        question_files = sorted(path for path in question.rglob("*") if path.is_file())
        if question_files:
            return question_files

    root_files = sorted(
        path
        for path in root.iterdir()
        if path.is_file() and path.suffix.lower() in SUPPORTED
    )
    if root_files:
        return root_files
    raise SystemExit(
        f"No supported problem inputs found in {question} or project root {root}"
    )


def main() -> None:
    parser = argparse.ArgumentParser(description="Inventory immutable question inputs.")
    parser.add_argument("project_root", type=Path)
    parser.add_argument("--refresh-baseline", action="store_true")
    args = parser.parse_args()
    root = args.project_root.resolve()

    entries = []
    for path in discover_input_files(root):
        if path.is_symlink():
            raise SystemExit(f"Symbolic links are not allowed in immutable input: {path}")
        details = inspect_file(path)
        entries.append({"path": path.relative_to(root).as_posix(), "size_bytes": path.stat().st_size, "sha256": sha256(path), "supported": path.suffix.lower() in SUPPORTED, "status": details["status"], "details": details})

    state = control_directory(root)
    state.mkdir(parents=True, exist_ok=True)
    inventory = state / "input-inventory.json"
    inventory.write_text(json.dumps({"schema_version": 1, "files": entries}, ensure_ascii=False, indent=2, default=json_default) + "\n", encoding="utf-8")
    baseline = state / "question-fingerprint.json"
    baseline_state = "preserved"
    if args.refresh_baseline or not baseline.exists():
        baseline.write_text(json.dumps({"schema_version": 1, "files": [{"path": item["path"], "sha256": item["sha256"]} for item in entries]}, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        baseline_state = "created" if not args.refresh_baseline else "refreshed explicitly"
    print(f"[OK] Inspected {len(entries)} question file(s): {inventory}")
    print(f"  Fingerprint baseline {baseline_state}: {baseline}")
    for item in entries:
        blank = item["details"].get("blank")
        note = " (blank; preserve only)" if blank else ""
        print(f"  [{item['status']}] {item['path']}: {item['size_bytes']} bytes{note}")


if __name__ == "__main__":
    configure_utf8_output()
    main()
